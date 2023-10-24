import os
import openpyxl
from dataclasses import dataclass
import pandas as pd
import datetime

@dataclass
class unitData:
    """Dataclass for moving unit data"""
    warnings: str = ''
    landUnits: pd.DataFrame | None = None
    airUnits: pd.DataFrame | None = None
    navalUnits: pd.DataFrame | None = None

class dataFileError(Exception):
    """Error class for datafile error"""
    pass

class ExcelReader:
    """Class for reading the data from Excel file"""

    # Information for checking that the file is correct
    correctSheets = {'Land', 'Air',  'Naval', 'Markers', 'Control', 'Headings Legend', 'Counter Colors', 'Latest CS', 'Copyright'}
    neededSheets = {'Land', 'Air', 'Naval'}
    requiredLandColumns = {'POWER', 'HOME', 'CLASS', 'TYPE', 'UNIT', 'YEAR', 'Strength', 'Réorg', 'Move', 'COST', 'TIME', 'SIZE', 'OTHER', 'KIT', 'OPTION', 'USED CLC', 'USED DLX'}
    requiredNavalColumns = {'SU', 'POWER', 'HOME', 'CLASS', 'TYPE', 'T2', 'SHIP1', 'SHIP2 / BACK / OTHER', 'YEAR', 'ATT', 'DEF', 'AA', 'SB', 'CV', 'MOV', 'RNG', 'COST1', 'COST2', 'TIME', 'SUNK', 'KIT', 'OPTION', 'USED2'}
    requiredAirColumns = {'POWER', 'HOME', 'CLASS', 'TYPE', 'UNIT', 'NAME', 'YEAR', 'ATA', 'ATS', 'TAC', 'STR', 'RANGE', '$(ifP)', '$(noP)', 'TIME', 'OTHER', 'KIT', 'OPTION', 'USED1', 'YR1', 'SIZ1', 'YR2', 'SIZ2', 'YR3', 'SIZ3'}
    correctModifier = 'Froon'
    correctDate = datetime.datetime(2019, 1, 1, 16, 12, 41)

    def openFile(self, path):
        """Opening the data file for reading
        
        Keyword arguments:
        path -- filepath to excel file
        """
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as err:
            raise dataFileError(f'Unable to open file {path}')

        if not self.neededSheets.issubset(wb.sheetnames):
            raise dataFileError(f'Excel-file does not have required sheets')

        # read and check land unit column names 
        land = wb["Land"]
        landColumnsGen = land.iter_rows(min_row=4, max_row=4, values_only=True)
        landColumns = next(landColumnsGen)

        if not self.requiredLandColumns.issubset(landColumns):
            raise dataFileError(f'Excel-file does not have required land unit data column labels on row 4')
        
        # read and check air unit column names
        air = wb["Air"]
        airColumnsGen = air.iter_rows(min_row=4, max_row=4, values_only=True)
        airColumns = next(airColumnsGen)

        if not self.requiredAirColumns.issubset(airColumns):
            raise dataFileError(f'Excel-file does not have required air unit data column labels on row 4')
        
        # read and check naval unit column names
        naval = wb["Naval"]
        navalColumnsGen = naval.iter_rows(min_row=4, max_row=4, values_only=True)
        navalColumns = next(navalColumnsGen)

        if not self.requiredNavalColumns.issubset(navalColumns):
            raise dataFileError(f'Excel-file does not have required naval unit data column labels on row 4')
        

        # create object to save unit data
        units = unitData()

        # create warning string for possible data issues
        if not wb.properties.lastModifiedBy == self.correctModifier:
            units.warnings+=(f'File creator not Froon, but {wb.properties.lastModifiedBy}, ')

        if not wb.properties.modified == self.correctDate:
            units.warnings+=(f'File last save date is not Jan 1st 2019, but {wb.properties.modified.isoformat()}, ')
        

        units.warnings.rstrip(', ')

        # read actual unit data to pandas and save them to data object
        units.airUnits=pd.read_excel(wb,sheet_name="Air", header=3, engine="openpyxl",usecols=self.requiredAirColumns)
        units.landUnits=pd.read_excel(wb,sheet_name="Land", header=3, engine="openpyxl",usecols=self.requiredLandColumns)
        units.navalUnits=pd.read_excel(wb,sheet_name="Naval", header=3, engine="openpyxl", usecols=self.requiredNavalColumns)

        return units